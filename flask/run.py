import datetime
import json
import uuid

import requests
from flask import Flask, jsonify, Response, request
from flask_admin import Admin
from flask_admin.contrib.sqla import ModelView
from flask_cors import CORS
from openpyxl import Workbook
from openpyxl.styles import Alignment

from models import *

# import time

app = Flask(__name__)
cors = CORS(app, resources={r"/api/*": {"origins": "*"}})  # 使前端能从后端拿到数据
app.config['SECRET_KEY'] = os.urandom(24)

admin = Admin(app, template_mode='bootstrap3')
admin.add_view(ModelView(Teacher, Session()))
admin.add_view(ModelView(Major, Session()))
admin.add_view(ModelView(ClassName, Session()))


@app.route("/api/GET/workday-data")
def get_workday_data():
    """
    api:获取工作日表
    """
    year = request.args.get("year", "2019")
    filename = year + 'workday_data.json'
    workday_data = dict()

    try:
        with open(filename) as file_obj:
            workday_data = json.load(file_obj)

    except FileNotFoundError:
        url = 'http://www.mxnzp.com/api/holiday/list/year/' + year
        request_api_text = requests.get(url).text
        request_api_json = json.loads(request_api_text)
        for each_month in request_api_json['data']:
            for each_day in each_month['days']:
                workday_data[each_day['date']] = each_day['typeDes'] == '工作日'
        with open(filename, 'w', encoding='utf8') as api_file_obj:
            json.dump(workday_data, api_file_obj, ensure_ascii=False)

    return jsonify({'isWorkdayData': workday_data})


########################################################################################################################

def get_columns_options_data():
    """
    api:获取表头以及对应列的讲师选项
    """
    session = Session()
    columns_data = []
    options_data = []

    for each_major in session.query(Major).filter_by(is_show=True).order_by(Major.order).all():
        # 表头
        columns_data.append({'title': each_major.title})
        # 对应列的讲师选项
        names = []
        for teacher in each_major.teacher:
            names.append(teacher.name)
        options_data.append(names)

    session.close()
    return columns_data, options_data


@app.route('/api/GET/table-data')
def get_table_data():
    session = Session()

    table_data = []
    for each_class in session.query(ClassName).filter_by(is_show=True).all():
        class_infos = []
        for each_major in session.query(Major).filter_by(is_show=True).order_by(Major.order).all():
            archive = session.query(Archive).filter_by(class_id=each_class.id, major_id=each_major.id).first()
            if archive:
                info = json.loads(archive.info)
            else:
                info = {
                    'startDate': None,
                    'endDate': None,
                    'duration': each_major.duration,
                    'teacher': None,
                    'conflictArray': [],
                    'id': each_major.id
                }
            class_infos.append(info)
        class_infos.append(each_class.class_name)
        class_infos.append(each_class.id)
        table_data.append(class_infos)

    session.close()
    columns_data, options_data = get_columns_options_data()
    return jsonify({
        'tableData': table_data,
        'columnsData': columns_data,
        'optionsData': options_data
    })


########################################################################################################################

def get_correct_date(date_str):
    correct_date = datetime.date(
        int(date_str[:4]),
        int(date_str[5:7]),
        int(date_str[8:10])
    ) + datetime.timedelta(1)

    return str(correct_date) + date_str[10:]


def get_xlsx(table_data, session):
    file_name = str(uuid.uuid4()) + '.xlsx'
    work_book = Workbook()  # 创建文件对象
    work_sheet = work_book.active  # 获取第一个sheet
    align = Alignment(horizontal='center', vertical='center')

    def _cell(row, col, _ws=work_sheet):
        return _ws.cell(row=row, column=col)

    def put_major():
        index = 2
        for each_major in session.query(Major).filter_by(is_show=True).order_by(Major.order).all():
            _cell(1, index).value = each_major.title
            _cell(1, index).alignment = align
            index += 1

    def put_class():
        index = 2
        for each_class in session.query(ClassName).filter_by(is_show=True).all():
            work_sheet.merge_cells(start_row=index, start_column=1, end_row=index + 3, end_column=1)
            _cell(index, 1).value = each_class.class_name
            _cell(index, 1).alignment = align
            index += 5

    row_i, col_i = 2, 2
    for row in table_data:
        for col in row:
            if isinstance(col, dict):
                _cell(row_i, col_i).value = '开始时间：%s' % get_correct_date(col['startDate'])[:10] \
                    if col['startDate'] else '开始时间：'
                _cell(row_i + 1, col_i).value = '结束时间：%s' % get_correct_date(col['endDate'])[:10] \
                    if col['endDate'] else '结束时间：'
                _cell(row_i + 2, col_i).value = '周    期：%s' % col['duration']
                _cell(row_i + 3, col_i).value = '讲    师：%s' % col['teacher']
                work_sheet.column_dimensions[_cell(row_i, col_i).column_letter].width = 25
            col_i += 1
        row_i += 5
        col_i = 2

    work_sheet.column_dimensions['A'].width = 12

    put_major()
    put_class()
    work_book.save(file_name)
    return file_name


@app.route('/api/POST/table-data', methods=['POST'])
def post_table_data():
    session = Session()

    table_data = request.json.get('tableData')

    for each_row in table_data:
        for each_col in each_row:
            if isinstance(each_col, dict):
                old_obj = session.query(Archive).filter_by(class_id=each_row[-1], major_id=each_col['id']).first()
                if old_obj:
                    session.delete(old_obj)
                    each_col['startDate'] = get_correct_date(each_col['startDate']) if each_col['startDate'] else ''
                    each_col['endDate'] = get_correct_date(each_col['endDate']) if each_col['endDate'] else ''

                session.add(
                    Archive(
                        class_id=each_row[-1],
                        major_id=each_col['id'],
                        info=json.dumps(each_col, ensure_ascii=False)
                    )
                )

    session.commit()

    response = Response(status=200)
    if request.json.get('isXlsx'):
        response = jsonify({'fileName': get_xlsx(table_data=table_data, session=session)})

    session.close()

    return response


@app.route('/xlsx/<file_name>')
def get_file_obj(file_name):
    try:
        file_obj = open(file_name, mode='rb')
        file_obj_r = file_obj.read()
        file_obj.close()
        os.remove(file_name)
        return Response(response=file_obj_r,
                        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    except FileNotFoundError:
        return Response(status=404)


########################################################################################################################

@app.route('/api/POST/new-class', methods=['POST'])
def post_new_class():
    new_class_name = request.json.get('newClassName')

    session = Session()
    new_class_obj = ClassName(class_name=new_class_name)
    session.add(new_class_obj)
    session.commit()

    new_class = []
    for each_major in session.query(Major).filter_by(is_show=True).order_by(Major.order).all():
        new_class.append(
            {
                'startDate': None,
                'endDate': None,
                'duration': each_major.duration,
                'teacher': None,
                'conflictArray': [],
                'id': each_major.id
            }
        )
    new_class.append(new_class_obj.class_name)
    new_class.append(new_class_obj.id)

    session.close()
    return jsonify({'newClass': new_class})


if __name__ == '__main__':
    app.run(host="0.0.0.0", port=5000, threaded=True)
