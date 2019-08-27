import re
from io import BytesIO
from zipfile import ZipFile, ZIP_DEFLATED

from django.http import HttpRequest, HttpResponse
from django.views import View
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.writer.excel import ExcelWriter

from schedule.views import save, deserialization


class XLSX(View):
    @staticmethod
    def post(request: HttpRequest):
        headers_and_options, row_header, schedules_data = deserialization(request)
        save(headers_and_options, row_header, schedules_data)

        xlsx_obj = XlsxFile(majors=headers_and_options, classes=row_header, schedules=schedules_data)
        xlsx_obj.fill()
        file_context = xlsx_obj.get_file()

        return HttpResponse(file_context)


class XlsxFile(object):
    align = Alignment(horizontal='center', vertical='center')

    def __init__(self, majors, classes, schedules):
        self.majors = majors
        self.classes = classes
        self.schedules = schedules
        self.work_book = Workbook()  # 创建文件对象
        self.work_sheet = self.work_book.active  # 获取第一个sheet

    def _cell(self, row, col):
        return self.work_sheet.cell(row=row, column=col)

    def put_major(self):
        index = 2
        for each_major in self.majors:
            self._cell(1, index).value = each_major['majorTitle']
            self._cell(1, index).alignment = self.align
            index += 1

    def put_class(self):
        index = 2
        for each_class in self.classes:
            self.work_sheet.merge_cells(start_row=index, start_column=1, end_row=index + 3, end_column=1)
            self._cell(index, 1).value = each_class['className']
            self._cell(index, 1).alignment = self.align
            index += 5

    def fill(self):
        row_i, col_i = 2, 2
        for row in self.schedules:
            for col in row:
                self._cell(row_i, col_i).value = '开始时间：{}'.format(self.format_date(col['startDate']))
                self._cell(row_i + 1, col_i).value = '结束时间：{}'.format(self.format_date(col['endDate']))
                self._cell(row_i + 2, col_i).value = '周    期：{}'.format(col['duration'])
                self._cell(row_i + 3, col_i).value = '讲    师：{}'.format(col['teacher'] or '')
                self.work_sheet.column_dimensions[self._cell(row_i, col_i).column_letter].width = 27
                col_i += 1
            row_i += 5
            col_i = 2

        self.work_sheet.column_dimensions['A'].width = 12

        self.put_major()
        self.put_class()

    def get_file(self):
        fp = BytesIO()
        archive = ZipFile(fp, 'w', ZIP_DEFLATED, allowZip64=True)
        writer = ExcelWriter(self.work_book, archive)
        writer.write_data()
        archive.close()
        fp.flush()
        fp.seek(0)
        return fp.read()

    @staticmethod
    def format_date(date_str):
        format_str = ''

        match = re.search(r'(\d{4})-(\d{2})-(\d{2})', str(date_str))
        if match:
            format_str = '{0[0]}年{0[1]}月{0[2]}日'.format(match.groups())

        return format_str
