import re
from os import remove
from uuid import uuid4

from flask import Blueprint, request, Response
from openpyxl import Workbook
from openpyxl.styles import Alignment

from .schedule import save

xlsx = Blueprint('xlsx', __name__)


@xlsx.route('/xlsx/', methods=['POST'])
def get_xlsx():
    headers_and_options = request.json.get('headersAndOptions')
    row_header = request.json.get('rowHeader')
    schedules_data = request.json.get('schedulesData')
    save(headers_and_options, row_header, schedules_data)

    xlsx_obj = XLSX(headers_and_options, row_header, schedules_data)
    xlsx_obj.fill()

    response = Response(
        response=xlsx_obj.get_file(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    return response


class XLSX(object):
    align = Alignment(horizontal='center', vertical='center')

    def __init__(self, majors, classes, schedules):
        self.majors = majors
        self.classes = classes
        self.schedules = schedules
        self.work_book = Workbook()  # 创建文件对象
        self.work_sheet = self.work_book.active  # 获取第一个sheet

    def _cell(self, row, col):
        return self.work_sheet.cell(row=row, column=col)

    def put_major(self):
        index = 2
        for each_major in self.majors:
            self._cell(1, index).value = each_major['title']
            self._cell(1, index).alignment = self.align
            index += 1

    def put_class(self):
        index = 2
        for each_class in self.classes:
            self.work_sheet.merge_cells(start_row=index, start_column=1, end_row=index + 3, end_column=1)
            self._cell(index, 1).value = each_class['className']
            self._cell(index, 1).alignment = self.align
            index += 5

    def fill(self):
        row_i, col_i = 2, 2
        for row in self.schedules:
            for col in row:
                self._cell(row_i, col_i).value = '开始时间：{}'.format(self.format_date(col['startDate']))
                self._cell(row_i + 1, col_i).value = '结束时间：{}'.format(self.format_date(col['endDate']))
                self._cell(row_i + 2, col_i).value = '周    期：{}'.format(col['duration'])
                self._cell(row_i + 3, col_i).value = '讲    师：{}'.format(col['teacher'] or '')
                self.work_sheet.column_dimensions[self._cell(row_i, col_i).column_letter].width = 27
                col_i += 1
            row_i += 5
            col_i = 2

        self.work_sheet.column_dimensions['A'].width = 12

        self.put_major()
        self.put_class()

    def get_file(self):
        file_name = '{}.xlsx'.format(str(uuid4()))
        self.work_book.save(file_name)
        with open(file_name, mode='rb') as fp:
            file_r = fp.read()
        remove(file_name)
        return file_r

    @staticmethod
    def format_date(date_str):
        format_str = ''

        match = re.search(r'(\d{4})-(\d{2})-(\d{2})', str(date_str))
        if match:
            format_str = '{0[0]}年{0[1]}月{0[2]}日'.format(match.groups())

        return format_str
